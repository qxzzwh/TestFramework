import inspect
import logging
import softest
from openpyxl import load_workbook

class Utils(softest.TestCase):
    def assertListItemText(self, list, value):
        for item in list:
            print("The text is: " + item.text)
            self.soft_assert(self.assertEqual, item.text, value)
            if item.text == value:
                print("assert pass")
            else:
                print("assert failed")
        self.assert_all()
    
    def custom_logger(self, loglevel=logging.DEBUG):
        # create logger
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(loglevel)
        # create console handler or file handler
        fh = logging.FileHandler("automation.log")
        # create formatter
        formatter1 = logging.Formatter(fmt='%(asctime)s - %(levelname)s - %(name)s : %(message)s', datefmt='%Y/%m/%d %I:%M:%S %p')
        # add formatter to console or file handler
        fh.setFormatter(formatter1)
        # add console handler to logger
        logger.addHandler(fh)

        return logger
    
    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column
        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(i,j).value)
            datalist.append(row)
        return datalist